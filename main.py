#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Print Server
Version: 2.3 - Fixed Printer Selection
"""

from flask import Flask, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl import Workbook
import os
from datetime import datetime
import tempfile
import sys

app = Flask(__name__)

# Настройки принтеров
PRINTERS = {
    8001: "MF270 Series",
    2: "HP LaserJet 1020", 
    3: "Epson L3150",
    4: "Brother HL-1110"
}

def create_excel_file(items, username=None, filial=None, order_id=None, category=None):
    """Создание Excel файла"""
    try:
        # Создание новой рабочей книги
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказ"
        
        # Настройка страницы для A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Не ограничивать по высоте
        
        # Настройка полей (в дюймах)
        ws.page_margins = PageMargins(
            left=0.2, right=0.2, top=0.2, bottom=0.2,
            header=0.3, footer=0.3
        )
        
        # Настройки стилей
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=11)
        total_font = Font(name='Arial', size=12, bold=True)
        info_font = Font(name='Arial', size=12, bold=True)
        
        # Цвета
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws['A1'] = 'СПИСОК ЗАКАЗОВ: ' + (order_id or 'N/A')
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        row = 3

        # Категория
        if category:
            ws[f'A{row}'] = 'Категория:  ' + category
            ws[f'A{row}'].font = info_font
            ws.merge_cells(f'A{row}:C{row}')
            row += 1

        # Заказчик
        if username:
            ws[f'A{row}'] = 'Заказчик:  ' + username
            ws[f'A{row}'].font = info_font
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
        
        # Филиал
        if filial:
            ws[f'A{row}'] = 'Филиал:  ' + filial
            ws[f'A{row}'].font = info_font
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
        
        # Время
        ws[f'A{row}'] = 'Время:   ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = info_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 2
        
        # Заголовок таблицы
        ws[f'A{row}'] = '№'
        ws[f'B{row}'] = 'Название товара'
        ws[f'C{row}'] = 'Ед. изм.'
        ws[f'D{row}'] = 'Кол-во'
        
        # Применение стиля заголовка
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row += 1
        start_data_row = row
        
        # Товары
        total_count = 0
        for i, item in enumerate(items, 1):
            product = item.get('name', item.get('product', 'Неизвестно'))
            count = item.get('count', 0)
            type_unit = item.get('type', 'шт')
            
            # Double sonlarni string yoki float sifatida saqlash
            try:
                count_value = float(count) if count else 0
            except (ValueError, TypeError):
                count_value = 0
            
            total_count += count_value
            
            ws[f'A{row}'] = i
            ws[f'B{row}'] = product
            ws[f'C{row}'] = type_unit
            ws[f'D{row}'] = count_value  # Float qiymat sifatida
            
            # Stili uchun number format o'rnatish
            ws[f'D{row}'].number_format = '0.00'  # Double sonlar uchun format
            
            # Стили для строк данных
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Название товара - крупный шрифт
            ws[f'B{row}'].font = Font(name='Arial', size=13.5, bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Единица измерения
            ws[f'C{row}'].font = normal_font
            ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Количество
            ws[f'D{row}'].font = normal_font
            ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Границы для всех ячеек
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = thin_border

            # Высота строки для лучшего отображения
            ws.row_dimensions[row].height = 25
            
            row += 1
        
        # Строка итого
        ws[f'A{row}'] = ''
        ws[f'B{row}'] = ''
        ws[f'C{row}'] = 'ИТОГО:'
        ws[f'D{row}'] = total_count
        
        # Number format uchun ИТОГО uchun ham
        ws[f'D{row}'].number_format = '0.00'
        
        # Применение стиля строки итого
        ws[f'C{row}'].font = total_font
        ws[f'D{row}'].font = total_font
        ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = total_fill
            ws[f'{col}{row}'].border = thin_border
        
        ws.row_dimensions[row].height = 29
        
        # Настройка ширины столбцов для A4 формата
        ws.column_dimensions['A'].width = 6      # Номер
        ws.column_dimensions['B'].width = 65     # Название товара
        ws.column_dimensions['C'].width = 10     # Единица измерения
        ws.column_dimensions['D'].width = 12     # Количество
        
        # Настройка печати
        ws.print_options.horizontalCentered = False
        ws.print_options.verticalCentered = False
        
        # Повторять заголовки на каждой странице
        ws.print_title_rows = f'1:{start_data_row-1}'
        
        # Имя файла и путь
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"zakaz_{timestamp}.xlsx"
        
        # Сохранение в папке temp
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        
        # Сохранение
        wb.save(filepath)
        
        print(f"✅ Excel файл создан: {filename}")
        return filepath, filename
        
    except Exception as e:
        print(f"❌ Ошибка при создании Excel: {e}")
        return None, None

def print_excel_file(filepath, printer_name):
    """Печать Excel файла на конкретный принтер"""
    try:
        if sys.platform != "win32":
            print("❌ Работает только на Windows")
            return False
        
        import win32print
        import win32api
        
        # Проверяем существование принтера
        printers = [printer[2] for printer in win32print.EnumPrinters(2)]
        print(f"🖨️  Sistemadagi barcha printerlar:")
        for idx, p in enumerate(printers, 1):
            print(f"     {idx}. {p}")
        
        print(f"\n🎯 Kerakli printer: '{printer_name}'")
        
        if printer_name not in printers:
            print(f"⚠️  OGOHLANTIRISH: Printer '{printer_name}' topilmadi!")
            print(f"   Mavjud printerlar ichidan qidirilmoqda...")
            
            # Попробуем найти похожий принтер
            found = False
            for p in printers:
                if printer_name.lower() in p.lower() or p.lower() in printer_name.lower():
                    print(f"   ✅ O'xshash printer topildi: {p}")
                    printer_name = p
                    found = True
                    break
            
            if not found:
                default = win32print.GetDefaultPrinter()
                print(f"   ⚠️  O'xshash printer topilmadi!")
                print(f"   🔄 Default printerga o'tkazilmoqda: {default}")
                printer_name = default
        else:
            print(f"✅ Printer topildi va tanlandi: {printer_name}")
        
        # USUL 1: Default printerni vaqtincha o'zgartirish
        print(f"\n🔄 USUL 1: Default printerni vaqtincha o'zgartirish...")
        
        # Hozirgi default printerni saqlash
        original_default = win32print.GetDefaultPrinter()
        print(f"   - Hozirgi default: {original_default}")
        
        try:
            # Yangi printerni default qilish
            win32print.SetDefaultPrinter(printer_name)
            print(f"   - Yangi default o'rnatildi: {printer_name}")
            
            # Excel orqali chop etish (agar Excel mavjud bo'lsa)
            print(f"\n📤 Chop etish...")
            print(f"   - Fayl: {filepath}")
            print(f"   - Printer: {printer_name}")
            
            excel_success = False
            try:
                import win32com.client
                
                print(f"   ✓ Excel orqali chop etish...")
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                
                wb = excel.Workbooks.Open(filepath)
                wb.PrintOut(Copies=1, Collate=True)
                wb.Close(False)
                excel.Quit()
                
                excel_success = True
                print(f"   ✅ Excel orqali chop etildi!")
                
            except Exception as excel_error:
                print(f"   ⚠️  Excel usuli ishlamadi: {excel_error}")
                print(f"   ✓ Standart usul bilan urinilmoqda...")
                
                # Standart usul
                win32api.ShellExecute(
                    0,
                    "print",
                    filepath,
                    None,
                    ".",
                    0
                )
                print(f"   ✅ Standart usul bilan yuborildi!")
            
            # Biroz kutish (chop etish jarayoni uchun)
            import time
            time.sleep(2)
            
        finally:
            # Default printerni qaytarish
            print(f"\n🔄 Default printerni qaytarish: {original_default}")
            win32print.SetDefaultPrinter(original_default)
            print(f"   ✅ Default printer qaytarildi")
        
        print(f"\n✅ Chop etish jarayoni tugallandi!")
        return True
            
    except ImportError as ie:
        print(f"❌ XATO: Kutubxona topilmadi: {ie}")
        print("💡 O'rnatish: pip install pywin32")
        return False
    except Exception as e:
        print(f"❌ XATO: Chop etishda xatolik: {e}")
        import traceback
        print("📋 Batafsil xatolik:")
        traceback.print_exc()
        
        # Default printerni qaytarishga harakat
        try:
            if 'original_default' in locals():
                win32print.SetDefaultPrinter(original_default)
                print(f"🔄 Default printer qaytarildi (xato bo'lishiga qaramay)")
        except:
            pass
        
        return False

@app.route('/', methods=['GET'])
def api_info():
    """Информация об API"""
    return jsonify({
        "name": "Excel Print Server",
        "version": "2.3",
        "status": "Работает",
        "description": "Создает и печатает Excel файлы с поддержкой выбора принтера",
        "printers": PRINTERS,
        "features": [
            "Создание Excel (.xlsx) файлов",
            "Профессиональные форматы",
            "A4 печать оптимизация",
            "Категории поддержка",
            "Единицы измерения (type)",
            "Выбор конкретного принтера",
            "Печать",
            "Скачивание",
            "Double числа поддержка"
        ],
        "endpoints": {
            "POST /print": "Создание и печать Excel",
            "POST /excel": "Только создание Excel (для скачивания)",
            "GET /printers": "Список доступных принтеров",
            "GET /": "Информация об API"
        },
        "example": {
            "url": "/print",
            "method": "POST",
            "body": {
                "printer": 8001,
                "username": "John Doe",
                "order_id": "25-09-12-10",
                "filial": "Ташкент",
                "category": "Продукты питания",
                "items": [
                    {
                        "product_id": 90,
                        "name": "Кортошка оллади",
                        "count": 4.5,
                        "type": "шт"
                    }
                ]
            }
        }
    })

@app.route('/printers', methods=['GET'])
def api_list_printers():
    """Список доступных принтеров в системе"""
    try:
        if sys.platform == "win32":
            import win32print
            printers = [printer[2] for printer in win32print.EnumPrinters(2)]
            default_printer = win32print.GetDefaultPrinter()
            
            return jsonify({
                "available_printers": printers,
                "default_printer": default_printer,
                "configured_printers": PRINTERS
            })
        else:
            return jsonify({
                "error": "Работает только на Windows",
                "configured_printers": PRINTERS
            })
    except ImportError:
        return jsonify({
            "error": "pywin32 не установлена",
            "configured_printers": PRINTERS
        })
    except Exception as e:
        return jsonify({
            "error": str(e),
            "configured_printers": PRINTERS
        })

@app.route('/print', methods=['POST'])
def api_print():
    """Создание и печать Excel"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Необходимы JSON данные"}), 400
        
        # ========== LOGGING: Kelgan ma'lumotlar ==========
        print("\n" + "="*60)
        print("📥 YANGI SO'ROV KELDI - /print endpoint")
        print("="*60)
        print(f"📦 Kelgan JSON data:")
        import json
        print(json.dumps(data, indent=2, ensure_ascii=False))
        print("-"*60)
        
        printer_key = data.get('printer', 1)
        items = data.get('items', [])
        username = data.get('username')
        filial = data.get('filial')
        order_id = data.get('order_id')
        category = data.get('category')
        
        # ========== LOGGING: Printer ma'lumotlari ==========
        print(f"🖨️  PRINTER MA'LUMOTLARI:")
        print(f"   - Kelgan printer key: {printer_key} (Type: {type(printer_key)})")
        print(f"   - Mavjud printer keys: {list(PRINTERS.keys())}")
        print(f"   - Printer key mavjudmi? {printer_key in PRINTERS}")
        
        if not items:
            print("❌ XATO: Mahsulotlar ro'yxati bo'sh!")
            return jsonify({"error": "Список товаров пуст"}), 400
        
        if printer_key not in PRINTERS:
            print(f"❌ XATO: Printer {printer_key} topilmadi!")
            return jsonify({"error": f"Принтер не найден. Доступные ключи: {list(PRINTERS.keys())}"}), 400
        
        printer_name = PRINTERS[printer_key]
        print(f"   - Tanlangan printer nomi: {printer_name}")
        print("-"*60)
        
        # Создание Excel файла
        print(f"📄 Excel fayl yaratilmoqda...")
        print(f"   - Order ID: {order_id}")
        print(f"   - Kategoriya: {category}")
        print(f"   - Mahsulotlar soni: {len(items)}")
        filepath, filename = create_excel_file(items, username, filial, order_id, category)
        if not filepath:
            print("❌ XATO: Excel fayl yaratilmadi!")
            return jsonify({"error": "Не удалось создать Excel"}), 500
        
        print(f"✅ Excel fayl yaratildi: {filename}")
        print(f"   - Fayl yo'li: {filepath}")
        print("-"*60)
        
        # Печать
        print(f"🖨️  CHOP ETISH BOSHLANDI:")
        print(f"   - Printer: {printer_name} (Key: {printer_key})")
        print_success = print_excel_file(filepath, printer_name)
        
        if print_success:
            print(f"✅ MUVAFFAQIYAT: Chop etish tayyor!")
        else:
            print(f"⚠️  OGOHLANTIRISH: Chop etishda muammo!")
        print("="*60 + "\n")
        
        response = {
            "success": True,
            "message": f"Excel файл создан и отправлен на {printer_name}",
            "filename": filename,
            "printer": printer_name,
            "printer_key": printer_key,
            "items_count": len(items),
            "total_quantity": sum(float(item.get('count', 0)) if item.get('count') else 0 for item in items),
            "print_status": "отправлен" if print_success else "ошибка печати",
            "file_path": filepath,
            "order_id": order_id,
            "category": category
        }
        
        return jsonify(response)
        
    except Exception as e:
        import traceback
        return jsonify({
            "error": f"Ошибка: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500

@app.route('/excel', methods=['POST'])
def api_excel_only():
    """Только создание Excel (для скачивания)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Необходимы JSON данные"}), 400
        
        items = data.get('items', [])
        username = data.get('username')
        filial = data.get('filial')
        order_id = data.get('order_id')
        category = data.get('category')
        
        if not items:
            return jsonify({"error": "Список товаров пуст"}), 400
        
        # Создание Excel файла
        filepath, filename = create_excel_file(items, username, filial, order_id, category)
        if not filepath:
            return jsonify({"error": "Не удалось создать Excel"}), 500
        
        # Отправка файла для скачивания
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"error": f"Ошибка: {str(e)}"}), 500

@app.route('/test', methods=['GET'])
def api_test():
    """Создание тестового Excel"""
    test_items = [
        {
            "product_id": 90,
            "name": "Кортошка оллади",
            "count": 4.5,
            "type": "шт"
        },
        {
            "product_id": 91,
            "name": "Молоко свежее",
            "count": 2.75,
            "type": "л"
        },
        {
            "product_id": 92,
            "name": "Хлеб белый",
            "count": 3.25,
            "type": "шт"
        },
        {
            "product_id": 93,
            "name": "Сахар",
            "count": 1.5,
            "type": "кг"
        },
        {
            "product_id": 94,
            "name": "Масло подсолнечное",
            "count": 1.2,
            "type": "л"
        }
    ]
    
    filepath, filename = create_excel_file(
        test_items, 
        "Тестовый пользователь", 
        "Тестовый филиал", 
        "TEST-001",
        "Продукты питания"
    )
    
    if filepath:
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        return jsonify({"error": "Не удалось создать тестовый Excel"}), 500

if __name__ == '__main__':
    print("=== EXCEL PRINT SERVER (FIXED PRINTER SELECTION) ===")
    print("URL: http://localhost:2020")
    print("Тестовый Excel: http://localhost:2020/test")
    print("Список принтеров: http://localhost:2020/printers")
    print("Принтеры:", list(PRINTERS.keys()))
    print("Excel библиотека: openpyxl")
    print("Новые функции: Правильный выбор принтера, Type/Unit support")
    print("Ctrl+C - Остановить")
    print("=====================================================")
    
    # Проверка библиотек
    try:
        import openpyxl
        print("✅ Библиотека openpyxl доступна")
    except ImportError:
        print("❌ Библиотека openpyxl отсутствует!")
        print("💡 Установка: pip install openpyxl")
    
    try:
        import win32print
        import win32api
        print("✅ Библиотека pywin32 доступна")
        
        # Показываем доступные принтеры
        printers = [printer[2] for printer in win32print.EnumPrinters(2)]
        default = win32print.GetDefaultPrinter()
        print(f"🖨️  Доступные принтеры в системе:")
        for p in printers:
            marker = " (DEFAULT)" if p == default else ""
            print(f"   - {p}{marker}")
    except ImportError:
        print("⚠️  Библиотека pywin32 отсутствует!")
        print("💡 Установка: pip install pywin32")
        print("   Без этой библиотеки печать будет идти на принтер по умолчанию")
    
    app.run(host='0.0.0.0', port=2020, debug=False)